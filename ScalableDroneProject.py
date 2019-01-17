#Author-Ajay(Your Father)
#Description-

import adsk.core, adsk.fusion, adsk.cam, traceback

#Basic core modules
import sys, os, inspect, subprocess, time


#Updated Basic modules
from .Modules import pip

#Modules needed to read Excel sheet
from .Modules import xlrd
#from .Modules import numpy as np
from .Modules import xlwt


script_path = os.path.abspath(inspect.getfile(inspect.currentframe()))
script_name = os.path.splitext(os.path.basename(script_path))[0]
script_dir = os.path.dirname(script_path)

sys.path.append(script_dir + "\Modules")
try:
    from openpyxl import load_workbook
    
    
finally:
    del sys.path[-1]
    
#from .Modules import win32com
#Moduels needed to write Excel Sheet
#from .Modules import openpyxl
#from .Modules.openpyxl import load_workbook
#from .Modules import editpyxl
#from .Modules import xlsxwriter
#from .Modules import xlutils
#from .Modules.xlutils import copy
#from .Modules import workbook
#from .Modules import worksheet


#Path of the Excel sheet(r: means raw address)
fileLocation = r"C:\Users\Internet\Desktop\Ajay's Project\Scalable Drone(Second Phase)\ScalableDroneProject(Spyder Script)\Mod UAV - Updated.xlsx"
sheetName = 'Initial Sizing'

#workbsook = xlrd.open_workbook(fileLocation)
workbookRead = xlrd.open_workbook(fileLocation, on_demand = True)
worksheetRead = workbookRead.sheet_by_name(sheetName)

workbookWrite = load_workbook(fileLocation)
worksheetWrite = workbookWrite.get_sheet_by_name(sheetName)
#worksheetWrite['A1'] = 'A1'
#workbookWrite.save(fileLocation)
#print(workbookWrite.sheetnames)
#sheetNamesAll = workbook.sheet_names()
#print(sheetNamesAll)

   
#Selecting which sheet to get the data
#workbook = xlutils.copy(workbook)

## get Sheet
#source = workbookWrite.get_sheet_by_name(sheetName)
#
## copy sheet
#target = workbookWrite.copy_worksheet(source)


#worksheetWrite = workbookWrite[sheetName]

# Creating classes to import datas
class Motor:
    
    def __init__(self, modelName, mass, baseDiameter, holeSize, holeDistanceXaxis, holeDistanceYaxis, maxCurrentDraw, recommendedVoltage):
        self.modelName = modelName
        self.mass = mass
        self.measurements = self.Measurements( baseDiameter, holeSize, holeDistanceXaxis, holeDistanceYaxis)
        self.maxCurrentDraw = maxCurrentDraw
        self.recommendedVoltage = recommendedVoltage

    def show(self):
        print("\n\n\n\nMotor Model: {0}\nMass: {1}g\nMax Current Draw: {2}A\nRecommended Voltage: {3}V".format( self.modelName, self.mass, self.maxCurrentDraw, self.recommendedVoltage))
        self.measurements.show()
        

        
    class Measurements:
        
        def __init__(self, baseDiameter, holeSize, holeDistanceXaxis, holeDistanceYaxis):
            self.baseDiameter = baseDiameter
            self.holeSize = holeSize
            self.holeDistanceXaxis = holeDistanceXaxis
            self.holeDistanceYaxis = holeDistanceYaxis

        def show(self):
            print("Base Diameter: {0}mm\nHole Size: {1}mm\nHole Distance X axis: {2}mm\nHole Distance Y axis: {3}mm\n".format( self.baseDiameter, self.holeSize, self.holeDistanceXaxis, self.holeDistanceYaxis))

class Camera:

    def __init__(self, modelName, mass, frameRate, resolution, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.frameRate = frameRate
        self.resolution = resolution
        self.measurements = self.Measurements( length, width, height)

    def show(self):
        print("\n\n\n\nCamera Model: {0}\nMass: {1}g\nFrame Rate: {2}fps\nResolution: {3}p".format( self.modelName, self.mass, self.frameRate, self.resolution))
        self.measurements.show()


    class Measurements:

        def __init__( self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm\n".format( self.length, self.width, self.height))

class FlightController:

    def __init__(self, modelName, mass, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.measurements = self.Measurements( length, width, height)

    def show(self):
        print("\n\n\n\nFCC Model: {0}\nMass: {1}g".format( self.modelName, self.mass))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm".format( self.length, self.width, self.height))
                
class ElectronicSpeedController:

    def __init__(self, modelName, massPerArm, length, width, height, currentFactor, maxCurrentRequired, typeESC, maxContinousCurrent, builtInESC):
        self.modelName = modelName
        self.massPerArm = massPerArm
        self.measurements = self.Measurements(length, width, height)
        self.currentFactor = currentFactor
        self.maxCurrentRequired = maxCurrentRequired
        self.typeESC = typeESC
        self.maxContinousCurrent = maxContinousCurrent
        self.builtInESC = builtInESC

    def show(self):
        print("\n\n\n\nESC Model: {0}\nMass Per Arm: {1}g\nCurrent Factor: {2}\nMax Current Required: {3}A\nESC Type: {4}\nMax Continous Current: {5}A\nBuilt-in ESC: {6}".format(self.modelName, self.massPerArm, self.currentFactor, self.maxCurrentRequired, self.typeESC, self.maxContinousCurrent, self.builtInESC))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm\n".format( self.length, self.width, self.height))       
        
class Battery:

    def __init__(self, modelName, mass, capacity, cell, discharge, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.capacity = capacity
        self.cell = cell
        self.discharge = discharge
        self.measurements = self.Measurements(length, width, height)

    def show(self):
        print("\n\n\n\nBattery Model: {0}\nMass: {1}g\nCapacity: {2}mAh\nCell: {3}S\nDischarge: {4}C".format(self.modelName, self.mass, self.capacity, self.cell, self.discharge))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height
        
        def show(self):
           print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm".format( self.length, self.width, self.height))         

class Telemetry:

    def __init__(self, modelName, mass, antennaLength, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.antennaLength = antennaLength
        self.measurements = self.Measurements(length, width, height)

    def show(self):
        print("\n\n\n\nTelemetry Model: {0}\nMass: {1}g\nAntenna Length: {2}mm".format(self.modelName, self.mass, self.antennaLength))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm".format( self.length, self.width, self.height))

class GPS:
    
    def __init__ (self, modelName, mass, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.measurements = self.Measurements(length, width, height)

    def show(self):
        print("\n\n\n\nGPS Model: {0}\nMass: {1}g".format(self.modelName, self.mass))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm\n".format( self.length, self.width, self.height))
   

class PowerDistributionBoard:
    
    def __init__(self, modelName, mass, length, width, height):
        self.modelName = modelName
        self.mass = mass
        self.measurements = self.Measurements(length, width, height)

    def show(self):
        print("\n\n\n\nPDB Model: {0}\nMass: {1}g".format(self.modelName, self.mass))
        self.measurements.show()

    class Measurements:

        def __init__(self, length, width, height):
            self.length = length
            self.width = width
            self.height = height

        def show(self):
            print("Length: {0}mm\nWidth: {1}mm\nHeight: {2}mm\n".format( self.length, self.width, self.height))

class Drone:

    def __init__(self, length, width, height, mass, centerGravity, area, density, volume, thickness, offsetLength, offsetWidth, offsetHeight):
        self.length = length
        self.width = width
        self.height = height
        self.mass = mass
        self.centerGravity = centerGravity
        self.area = area
        self.density = density
        self.volume = volume
        self.design = self.Design(thickness, offsetLength, offsetWidth, offsetHeight)
        

    class Design:
        
        def __init__(self, thickness, offsetLength, offsetWidth, offsetHeight):
            self.thickness = thickness
            self.offsetLength = offsetLength
            self.offsetWidth = offsetWidth
            self.offsetHeight = offsetHeight



        def show(self):
            print("\n\n\n\nDrone Design:\nThickness: {0}mm\nOffset Length: {1}\nOffset Width: {2}\nOffset Height: {3}".format(self.thickness, self.offsetLength, self.offsetWidth, self.offsetHeight))
class Propeller:
    
    def __init__(self, modelName, radius, mass):
        self.modelName = modelName
        self.radius = radius
        self.mass = mass
        
    def show(self):
        print("\n\n\n\nPropeller:\nModel Name: {0}\nRadius: {1}mm\nMass: {2}g\n".format(self.modelName, self.radius, self.mass))
        
        
class PrintSpecs:

    def __init__ (self, maxLength, maxWidth, maxHeight):
        self.maxLength = maxLength
        self.maxWidth = maxWidth
        self.maxHeight = maxHeight

    def show(self):
         print("\n\n\n\nPriner Bed Size:\nMax Length: {0}mm\nMax Width: {1}mm\nMax Height: {2}mm\n".format(self.maxLength, self.maxWidth, self.maxHeight))

class UserInput:
    
    def __init__(self, resolution, cameraStability, endurance, camMassAsMTOW):
        self.resolution = resolution
        self.cameraStability = cameraStability
        self.endurance = endurance
        self.camMassAsMTOW = camMassAsMTOW
        
    def show(self):
        print("User Input Selection:\nCamera Resolution selected: {0}p\nCamera Stability selected: {1}\nEndurance selected(hrs): {2}hrs\nCam-Mass-as-MTOW selected(%): {3}%\n".format(self.resolution, self.cameraStability, self.endurance, self.camMassAsMTOW))
    
#Function to open excel for User
def updateExcelSheet():
    
    #Getting the updated changes
    #Open Excel File
    def openExcel():
        os.startfile(fileLocation)

    def closeExcel():
        os.system('TASKKILL /F /IM EXCEl.exe')
        
    print("Please Save and Close excel sheet after selecting.")

    openExcel()
    
    def process_exists(process_name):
        call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
        # use buildin check_output right away
        output = subprocess.check_output(call)
#        # check in last line for process name
        last_line = output.strip().split()[-1]
            
        #print(last_line)
        if(last_line == b'K'):
            return True
        else:
            return False
#        # because Fail message could be translated
        #return last_line.lower().startswith(process_name).lower()

    print(process_exists('EXCEL.exe'))    
    
    
    while(process_exists('EXCEL.exe') == True):
        time.sleep(2.5)
        
    workbookRead = xlrd.open_workbook(fileLocation, on_demand = True)
    worksheetRead = workbookRead.sheet_by_name(sheetName)   
    
    
##Default Values
#Motor Properties            
motor = Motor('1105-7500KV', 5.9, 27.5, 3, 16, 19, 0, 0)

#Camera Properties
camera = Camera('GWY CMOS Camera with VCR', 20, 60, 720, 26, 28, 25)

#Flight Controller Properties
flightCC = FlightController('PixRacer R15', 10.54, 36, 36 , 11.5)

#Electronic Speed Controller Properties
esc = ElectronicSpeedController('Turnighy MultiStar Race 4-in-1 10A', 13, 27, 27, 3, 0, 0, 0, 0, 0)

#Battery Properties
battery = Battery('ZIPPY Compact', 264, 3700, 3, 25, 146, 43, 19)

#Telemetry Properties
telemetry = Telemetry('Holybro 100mW FPV Transceiver Telemetry Radio Set (433Mhz)', 'unknown', 26, 56, 26, 19)

#GPS Properties
gps = GPS('Ublox Neo-M9N', 22.4, 3.7,  3.7, 1.2)

#Power Distribution board Properties
pdb = PowerDistributionBoard('Diatone v8.3 LC Filtered', 5.5, 3.6,  3.6, 1)

#Drone Properties
drone = Drone(0, 0, 0, 0, 0, 0, 0, 0, 3, 10, 0, 5)

#Propeller Properties
prop = Propeller('GF5152-3' , 64.77, 5.3)

#3D Printer Properties
threeDprinter = PrintSpecs(215, 197, 200)

#User Input
userInput = UserInput(0, 0, 0, 0)

def droneLengthCalculation():
    
    totalLength = 0
    #totalLength = camera.measurements.length + flightCC.measurements.length + esc.measurements.length + telemetry.measurements.length + telemetry.antennaLength + (4*drone.design.offsetLength)

    #Second Method
    totalLengthArray = [camera.measurements.length, flightCC.measurements.length, esc.measurements.length, telemetry.measurements.length, telemetry.antennaLength, drone.design.offsetLength]
    
    for x in totalLengthArray:
        #print(x)
        
        if(x == drone.design.offsetLength):
            numberOfComponent = len(totalLengthArray) - 2
            totalLength = totalLength + (numberOfComponent)*(drone.design.offsetLength)
        elif(x != 0):
            totalLength = totalLength + x
        else:
            i = 0
            i = i + 1
            print("Warning: {0} is NULL in totalLengthArray list".format(i))
            
    
    return totalLength

def droneWidthCalculation():
    totalWidth = 0
    
    totalWidthArray = [camera.measurements.width, flightCC.measurements.width, esc.measurements.width, telemetry.measurements.width, battery.measurements.width]

    #print(totalWidthArray)

    totalWidthArray.sort(reverse = True)
    #print(totalWidthArray)
    totalWidth = totalWidth + totalWidthArray[0]

    for x in totalWidthArray:
        #print(x)

        if(x == 0):
            i = 0
            i = i + 0
            print("Warning: {0} is NULL in totalWidthArray list".format(i))



    totalWidth = totalWidth + 2*drone.design.offsetWidth
    
    return totalWidth

def droneHeightCalculation():
    totalHeight = 0

    totalHeightArray = [camera.measurements.height, flightCC.measurements.width, esc.measurements.height, telemetry.measurements.height]
    #print(totalHeightArray)
    totalHeightArray.sort(reverse = True)
    #print(totalHeightArray)

    totalHeight = totalHeight + totalHeightArray[0]

    totalHeight = totalHeight + drone.design.offsetHeight

    return totalHeight
    
#Function: makeing sure that data is not empyty
def checkingDataExcel(row, col):

    workbookRead = xlrd.open_workbook(fileLocation, on_demand = True)
    worksheetRead = workbookRead.sheet_by_name(sheetName) 

    if(worksheetRead.cell(row, col).value != xlrd.empty_cell.value):
        #print(workbookMainData.cell(axisY, axisX).value)
        return worksheetRead.cell(row, col).value
    else:
        print("Warning: Empty input at {0}\n({1},{2})".format(worksheetRead.cell(row, col).value, row, col))

#Search string cell function, returns value of the string 
def execlSearchCell(string, addRow, addCol):
    
    workbookRead = xlrd.open_workbook(fileLocation, on_demand = True)
    worksheetRead = workbookRead.sheet_by_name(sheetName)    
    
    i=0
    area = worksheetRead.nrows*worksheetRead.ncols
    for row in range(worksheetRead.nrows):
        for col in range(worksheetRead.ncols):
            currentCell = worksheetRead.cell(row, col)
            if(currentCell.value == string):
                return checkingDataExcel(row + addRow, col + addCol)
            elif(i + 1 == area):
                #print(i)
                print("\nInvalid string in Excel Sheet: {0}".format(string))
            else:
                i = i + 1

#Editing date from excel sheet
def editDataFrameExcel():
    
    workbookWrite = load_workbook(fileLocation)
    worksheetWrite = workbookWrite.get_sheet_by_name(sheetName)
    
    #fileGeneratedpath = r"C:\Users\Internet\Desktop\Ajay's Project\Scalable Drone(Second Phase)\ScalableDroneProject(Spyder Script)/python.xlsx"

    #save workbook
    #workbookWrite.save(fileGeneratedpath)

#    workbookWrite1 = openpyxl.load_workbook(fileGeneratedpath)
#    worksheetWrite1 = workbookWrite1[sheetName]
    
#    print(worksheetWrite.cell(6 +1,1 +1).value)
#    print(worksheetRead.cell(6,1).value)
#    worksheetWrite.cell(6 +1,1 +1).value = 1440
#    print(worksheetWrite.cell(6 +1,1 +1).value)
#    print(worksheetRead.cell(6,1).value)
    
    #Writing user inputs

    userInput.resolution = 720
    userInput.cameraStability = 0
    userInput.endurance = 0.3
    userInput.camMassAsMTOW = 0.1
    
    worksheetWrite['B7'] = userInput.resolution
    worksheetWrite['B8'] = userInput.cameraStability
    worksheetWrite['B9'] = userInput.endurance
    worksheetWrite['B15'] = userInput.camMassAsMTOW
    
    workbookWrite.save(fileLocation)
    

#Getting data from excel sheet
def getDataFrameExcel():
       
    convertFromMToMM = 1000
    convertToPercentage = 100
    addRow = 0
    #Getting the value back beside the search string
    addCol = 1
   
    userInput.resolution = execlSearchCell('Camera Resolution', addRow, addCol)
    userInput.cameraStability = execlSearchCell('Camera Stability (0 - min OR 3 - max)', addRow, addCol)
    userInput.endurance = execlSearchCell('Desired Endurance', addRow, addCol)
    userInput.camMassAsMTOW = execlSearchCell('Cam-mass-as-MTOW %', addRow, addCol)*convertToPercentage
    
    motor.modelName = execlSearchCell('Motor Model', addRow, addCol)
    motor.mass = execlSearchCell('Motor Mass', addRow, addCol)
    motor.maxCurrentDraw = execlSearchCell('Motor Max Current Draw', addRow, addCol)
    motor.recommendedVoltage = execlSearchCell('Motor Recommended Voltage', addRow, addCol)             

    camera.modelName = execlSearchCell('Camera Model', addRow, addCol)
    camera.resolution = execlSearchCell('Camera Resolution', addRow, addCol)
    camera.mass = execlSearchCell('Camera Mass', addRow, addCol)
    camera.frameRate = ('NIL ', addRow, addCol)
    
    flightCC.mass = execlSearchCell('FCC Mass', addRow, addCol)
    flightCC.measurements.length = round(execlSearchCell('FCC Length', addRow, addCol)*convertFromMToMM, 2)
    flightCC.measurements.width = round(execlSearchCell('FCC Width', addRow, addCol)*convertFromMToMM, 2)
    #flightCC.measurements.height = execlSearchCell('FCC Height')*convertToMM
    
    esc.modelName = execlSearchCell('ESC Model', addRow, addCol)
    esc.massPerArm = execlSearchCell('ESC Mass Per Arm', addRow, addCol)
    esc.currentFactor = execlSearchCell('ESC Current Factor', addRow, addCol)
    esc.maxCurrentRequired = round(execlSearchCell('ESC Current Required (Max)', addRow, addCol), 2)
    esc.typeESC = execlSearchCell('ESC Type', addRow, addCol)
    esc.maxContinousCurrent = execlSearchCell('ESC Max Continous Current', addRow, addCol)
    esc.builtInESC = execlSearchCell('ESC Built-in ESC', addRow, addCol)
    
    battery.modelName = execlSearchCell('Battery Model', addRow, addCol)
    battery.mass = execlSearchCell('Battery Mass', addRow, addCol)
    battery.capacity = execlSearchCell('Battery Capacity', addRow, addCol)
    battery.cell = execlSearchCell('Battery Cell Count', addRow, addCol)
    battery.measurements.length = round(execlSearchCell('Battery Length', addRow, addCol)*convertFromMToMM, 2)
    battery.measurements.width = round(execlSearchCell('Battery Width', addRow, addCol)*convertFromMToMM, 2)
    battery.measurements.height = round(execlSearchCell('Battery Height', addRow, addCol)*convertFromMToMM, 2)
    
    
def droneBody():
    drone.length = round(droneLengthCalculation(), 2)
    
    drone.width = round(droneWidthCalculation(), 2)
   
    drone.height = round(droneHeightCalculation(), 2)
    
    
    if(drone.length > threeDprinter.maxLength):
        print("Warning: Drone Length exceed 3D printer bed size.\nDrone Length:{0}mm\n3D printer Max Length:{1}mm".format(drone.length, threeDprinter.maxLength))
    else:
        print("Total Drone Length: {0}mm".format(drone.length))
        
    if(drone.width > threeDprinter.maxWidth):
        print("Warning: Drone Width exceed 3D printer bed size.\nDrone Width:{0}mm\n3D printer Max Width:{1}mm".format(drone.width, threeDprinter.maxWidth))
    else:
        print("Total Drone Width: {0}mm".format(drone.width))
        
    if(drone.height > threeDprinter.maxHeight):
         print("Warning: Drone Height exceed 3D printer bed size.\nDrone Height:{0}mm\n3D printer Max Height:{1}mm".format(drone.width, threeDprinter.maxWidth))
    else:
        print("Total Drone Height: {0}mm\n".format(drone.height))

def updateParameters():
    
    app = adsk.core.Application.get()
    ui  = app.userInterface    
    design = adsk.fusion.Design.cast(app.activeProduct)
    

    # Get the root component of the active design.
    rootComp = design.rootComponent
    #get array of occurrences
    perviousOccurrences = rootComp.occurrences.asList
    
    #Delete all User parameters
#    perviousParameters = design.userParameters
#    print(perviousParameters.count)
#    if(perviousParameters.count-1 > 0):
#        for x in range(perviousParameters.count-1):
#            perviousParameters.item(x+2).deleteMe()

            
    
    #Delete all previous Component
#    if(perviousOccurrences.count > 0):
#        for i in range(perviousOccurrences.count):
#            perviousOccurrences.item(i).deleteMe()



    # Access the import manager and root component
    importManager = app.importManager
    rootComp = app.activeProduct.rootComponent
            
    # Get the file to be imported, here we telling it where the file is located
    filename = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Drone Frame.f3d")
   
    # Create the input options and import them to the target
    importOptions = importManager.createFusionArchiveImportOptions(filename)
    importManager.importToTarget(importOptions, rootComp)    
    
    # Get the occurance of the imported component
    droneOccurance = rootComp.occurrences.item(rootComp.occurrences.count-1)
    
    # Accessing the parameters
    parameters = droneOccurance.component.parentDesign.allParameters
    parameters.itemByName('Drone_Length').expression = str(drone.length)
    parameters.itemByName('Drone_Width').expression = str(drone.width)
    parameters.itemByName('Drone_Height').expression = str(drone.design.thickness)
    
    
    
    #First method   
#    design = adsk.fusion.Design.cast(app.activeProduct)
#    userParams = design.userParameters
#        
#    userParams.itemByName('Drone_Length').expression = str(drone.length)
#    userParams.itemByName('Drone_Width').expression = str(drone.width)
#    userParams.itemByName('Drone_Height').expression = str(drone.height) 

def droneGeneratedData():
    
    app = adsk.core.Application.get()
    ui  = app.userInterface
    
    product = app.activeProduct
    design = adsk.fusion.Design.cast(product)    
    
    # Get the root component of the active design.
    rootComp = design.rootComponent
      
    # Get physical properties from component (very high accuracy)
    physicalProperties = rootComp.getPhysicalProperties(adsk.fusion.CalculationAccuracy.VeryHighCalculationAccuracy)
    
    #converting from g/mm^3 to g/cm^3
    convertDensity = 1000
    #converting from kg to g
    convertMass = 1000
    
    # Get data from physical properties
    drone.area = round(physicalProperties.area, 2)
    drone.density = round(physicalProperties.density*convertDensity, 2)
    drone.mass = round(physicalProperties.mass*convertMass, 2)
    drone.volume = round(physicalProperties.volume, 2)

    # Get center of mass from physical properties
    drone.centerGravity = physicalProperties.centerOfMass        
    
    print('Drone Generated:\nSurface Area: {0}mm^2\nDensity: {1}g/cm^3\nMass: {2}g\nVoulme: {3}cm^3\n'.format(drone.area, drone.density, drone.mass, drone.volume))
    
    
 
#Main Programmes
def run(context):
    ui = None
    try:
        app = adsk.core.Application.get()
        ui  = app.userInterface
        #ui.messageBox('Hello script')

        # Create a document.
        doc = app.documents.add(adsk.core.DocumentTypes.FusionDesignDocumentType)
        
        product = app.activeProduct
        design = adsk.fusion.Design.cast(product)
        
        # Get the root component of the active design.
        rootComp = design.rootComponent

        
        
        editDataFrameExcel()
        updateExcelSheet()
        getDataFrameExcel()
        droneBody()
        #Print default values
#        motor.show()
#        camera.show()
#        flightCC.show()
#        esc.show()
#        battery.show()
#        telemetry.show()
#        gps.show()
#        pdb.show()
#        drone.design.show()
#        prop.show()
#        threeDprinter.show()
#       userInput.show()
        
        updateParameters()
        droneGeneratedData()
        
        
        print('Hello World')
       

    except:
        if ui:
            ui.messageBox('Failed:\n{}'.format(traceback.format_exc()))
   
    
    
    
    